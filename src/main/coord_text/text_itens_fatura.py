import pdfplumber
import re
import os
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def extrair_texto_por_linhas(pdf_path, coordenadas, pagina=0):
    """Extrai texto por linhas de uma área específica do PDF"""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            pagina_pdf = pdf.pages[pagina]
            area = (
                coordenadas[0][0],
                coordenadas[0][1],
                coordenadas[1][0],
                coordenadas[1][1]
            )
            linhas = pagina_pdf.within_bbox(area).extract_text_lines()
            return linhas
    except Exception as e:
        print(f"Erro ao processar {pdf_path}: {e}")
        return []


def extrair_valores_apos_unidade(texto, unidades):
    """Extrai valores após as unidades KWH, UN ou KW"""
    for unidade in unidades:
        padrao = rf".*?{unidade}(.*)"
        match = re.search(padrao, texto, re.IGNORECASE)
        if match:
            parte_numerica = match.group(1).strip()
            return re.findall(r"-?\d{1,3}(?:\.\d{3})*(?:,\d+)?|-?\d+", parte_numerica)
    return []


def processar_pdf(pdf_path, coordenadas):
    """Processa um único PDF e retorna uma lista de dicionários com os dados"""
    dados_pdf = []

    linhas = extrair_texto_por_linhas(str(pdf_path), coordenadas)

    if not linhas:
        return dados_pdf

    for linha in linhas:
        texto_linha = linha['text']

        # Padrões de busca com unidades
        padroes = [
            (r'(Consumo.*?(?-i:KWH))', "Consumo", ["KWH"]),
            (r'(Energia.*?(?:KWH|UN|KW))', "Energia", ["KWH", "UN", "KW"]),
            (r'(Demanda.*?KW)', "Demanda", ["KW"]),
            (r'(Adic\. B\.)', "Adic. B.", []),
            (r'(Custo de Disponibilidade)', "Custo de Disponibilidade", []),
            (r'(.*TUSD[^\d]*(?:\d{2}/\d{4})?)', "TUSD", []),
            (r'(Ilum Pub)', "Ilum Pub", []),
            (r'(MULTA.*?\d{2}/\d{4})', "MULTA", []),
            (r'(JUROS DE.*?\d{2}/\d{4})', "JUROS DE", []),
            (r'(ATUALIZAÇÃO .*?\d{2}/\d{4})', "ATUALIZAÇÃO", []),
            (r'(PARCELA.*?\d{2}/\d{4})', "PARCELA", []),
            (r'(COMPENSACAO.*?\d{2}/\d{4})', "COMPENSACAO", []),
            (r'(DIF\.CREDITO.*?)(?=\d{2}/\d{4})', "DIF.CREDITO", []),
            (r'(Substituição.*?-(?: Crédito| Débito))', "Substituição", [])
        ]

        for padrao, tipo, unidades in padroes:
            m = re.search(padrao, texto_linha, re.IGNORECASE)
            if m:
                descricao = m.group(1).strip()

                # Extrair valores
                if unidades:
                    valores = extrair_valores_apos_unidade(texto_linha, unidades)
                else:
                    valores = re.findall(r"-?\d{1,3}(?:\.\d{3})*(?:,\d+)?|-?\d+", texto_linha)

                # Criar dicionário com os dados
                item_data = {
                    'arquivo': pdf_path.name,
                    'caminho': str(pdf_path),
                    'descricao': descricao
                }

                # Preencher os valores nas colunas corretas
                if len(valores) >= 8:
                    item_data.update({
                        'quant': valores[0],
                        'preco_unit_com_tributos': valores[1],
                        'valor': valores[2],
                        'pis_confins': valores[3],
                        'base_calc_icms': valores[4],
                        'porcent_icms': valores[5],
                        'icms': valores[6],
                        'tarifa_unit': valores[7]
                    })
                elif len(valores) == 5:
                    item_data.update({
                        'valor': valores[0],
                        'pis_confins': valores[1],
                        'base_calc_icms': valores[2],
                        'porcent_icms': valores[3],
                        'icms': valores[4]
                    })
                elif len(valores) == 1:
                    item_data.update({
                        'valor': valores[0]
                    })
                elif valores:
                    # Para outros casos, colocar o primeiro valor na coluna 'valor'
                    item_data.update({
                        'valor': valores[0]
                    })

                dados_pdf.append(item_data)
                break

    return dados_pdf


def processar_pasta_pdfs(pasta, coordenadas, output_xlsx="resultado.xlsx"):
    """Processa todos os PDFs de uma pasta e gera um XLSX único"""

    pdf_files = list(Path(pasta).glob("*.pdf"))

    if not pdf_files:
        print(f"Nenhum arquivo PDF encontrado na pasta: {pasta}")
        return

    print(f"Encontrados {len(pdf_files)} arquivos PDF para processar")

    todos_dados = []

    for i, pdf_path in enumerate(pdf_files, 1):
        print(f"Processando ({i}/{len(pdf_files)}): {pdf_path.name}")

        dados = processar_pdf(pdf_path, coordenadas)
        todos_dados.extend(dados)

    # Converter para DataFrame
    df = pd.DataFrame(todos_dados)

    # Definir ordem das colunas
    colunas_ordenadas = [
        'arquivo', 'caminho', 'descricao', 'quant', 'preco_unit_com_tributos',
        'valor', 'pis_confins', 'base_calc_icms', 'porcent_icms', 'icms', 'tarifa_unit'
    ]

    # Reordenar colunas
    df = df.reindex(columns=[col for col in colunas_ordenadas if col in df.columns])

    # Salvar como XLSX com formatação
    with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Dados', index=False)

        # Formatar a planilha
        workbook = writer.book
        worksheet = writer.sheets['Dados']

        # Formatar cabeçalho
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_num, value in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font

        # Ajustar largura das colunas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    print(f"Arquivo XLSX salvo como: {output_xlsx}")
    return df


# Configurações
pasta_pdfs = r"T:\vitor energia\FATURAS AGRICOLA 2025\2025\TODAS"  # Altere para o caminho da sua pasta
coordenadas = [(21.7, 361.2), (444.1, 571.7)]
output_xlsx = "conta_energia_padrao.xlsx"

# Processar todos os PDFs da pasta
df_resultado = processar_pasta_pdfs(pasta_pdfs, coordenadas, output_xlsx)

# Mostrar preview dos dados
print("\nPreview dos dados:")
print("=" * 80)
print(f"Total de itens processados: {len(df_resultado)}")
print(f"Colunas: {list(df_resultado.columns)}")
print(f"\nPrimeiras 5 linhas:")
print(df_resultado.head())