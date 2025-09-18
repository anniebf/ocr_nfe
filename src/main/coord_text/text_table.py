import pdfplumber
import re
import os
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

'''
O CODIGO PEGA AS COORDENDAS DO ITENS DA FATURA E RETORNA UM XLSX 
COM AS COLUNAS SENDO DESCRIÇÃO + VALOR E LINHAS SENDO OS ARQUIVOS PDF
'''


def extrair_texto_por_linhas(pdf_path, coordenadas, pagina=0):
    """Extrai texto por linhas de uma área específica do PDF"""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            pagina_pdf = pdf.pages[pagina]

            area = (
                coordenadas[0][0],
                coordenadas[0][1],
                coordenadas[1][0],
                coordenadas[1][1]
            )

            linhas = pagina_pdf.within_bbox(area).extract_text_lines()
            return linhas

    except Exception as e:
        print(f"Erro ao processar {pdf_path}: {e}")
        return []


def extrair_todos_valores(texto):
    """Extrai todos os valores numéricos, incluindo números simples"""
    return re.findall(r"-?\d{1,3}(?:\.\d{3})*(?:,\d+)?|-?\d+", texto)


def processar_pdf(pdf_path, coordenadas):
    """Processa um único PDF e retorna um dicionário com os dados"""
    dados_pdf = {
        'arquivo': pdf_path.name,
        'caminho': str(pdf_path)
    }

    linhas = extrair_texto_por_linhas(str(pdf_path), coordenadas)

    if not linhas:
        return dados_pdf

    for linha in linhas:
        texto_linha = linha['text']
        valores = extrair_todos_valores(texto_linha)

        # Padrões de busca
        padroes = [
            (r'(Consumo.*?(?-i:KWH))', "Consumo"),
            (r'(Custo.*?KWH)', "Custo"),
            (r'(Energia.*?(?:KWH|UN))', "Energia"),
            (r'(Demanda.*?KW)', "Demanda"),
            (r'(Adic\. B\.)', "Adic. B."),
            (r'(.*TUSD[^\d]*(?:\d{2}/\d{4})?)', "TUSD"),
            (r'(Ilum Pub)', "Ilum Pub"),
            (r'(JUROS DE.*?\d{2}/\d{4})', "JUROS DE"),
            (r'(MULTA.*?\d{2}/\d{4})', "MULTA"),
            (r'(ATUALIZAÇÃO .*?\d{2}/\d{4})', "ATUALIZAÇÃO "),
            (r'(PARCELA.*?\d{2}/\d{4})', "PARCELA"),
            (r'(Adicional [^\d]*)', "Adicional "),
            (r'(Substituição.*?-(?: Crédito| Débito))', "Substituição"),
            (r'(COMPENSACAO.*?\d{2}/\d{4})', "COMPENSACAO"),
            (r'(DIF\.CREDITO.*?)(?=\d{2}/\d{4})', "DIF.CREDITO"),
            (r'(COMPENSACAO.*?)(?=\d{2}/\d{4})', "COMPENSACAO")
        ]

        for padrao, tipo in padroes:
            m = re.search(padrao, texto_linha, re.IGNORECASE)
            if m:
                descricao = m.group(1).strip()

                # Determinar o valor
                if valores:
                    if tipo in ["DIF.CREDITO", "COMPENSACAO"]:
                        valor = valores[-1] if valores else ""
                    elif len(valores) >= 5:
                        valor = valores[0]  # Pega o primeiro valor para a maioria dos casos
                    else:
                        valor = valores[0] if valores else ""
                else:
                    valor = ""

                # Adicionar ao dicionário
                dados_pdf[descricao] = valor
                break

    return dados_pdf


def processar_pasta_pdfs(pasta, coordenadas, output_xlsx="resultado.xlsx"):
    """Processa todos os PDFs de uma pasta e gera um XLSX único"""

    # Encontrar todos os arquivos PDF na pasta
    pdf_files = list(Path(pasta).glob("*.pdf"))

    if not pdf_files:
        print(f"Nenhum arquivo PDF encontrado na pasta: {pasta}")
        return

    print(f"Encontrados {len(pdf_files)} arquivos PDF para processar")

    # Lista para armazenar todos os dados
    todos_dados = []
    todas_colunas = set(['arquivo', 'caminho'])  # Colunas base

    # Processar cada PDF
    for i, pdf_path in enumerate(pdf_files, 1):
        print(f"Processando ({i}/{len(pdf_files)}): {pdf_path.name}")

        dados = processar_pdf(pdf_path, coordenadas)
        todos_dados.append(dados)

        # Adicionar novas colunas encontradas
        todas_colunas.update(dados.keys())

    # Converter para DataFrame
    df = pd.DataFrame(todos_dados)

    # Reordenar colunas: arquivo, caminho, depois as demais em ordem alfabética
    colunas_ordenadas = ['arquivo', 'caminho'] + sorted(
        [col for col in todas_colunas if col not in ['arquivo', 'caminho']])
    df = df.reindex(columns=colunas_ordenadas)

    # Salvar como XLSX com formatação
    with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Dados', index=False)

        # Formatar a planilha
        workbook = writer.book
        worksheet = writer.sheets['Dados']

        # Formatar cabeçalho
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_num, value in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font

        # Ajustar largura das colunas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    print(f"Arquivo XLSX salvo como: {output_xlsx}")
    return df


# Configurações
pasta_pdfs = r"T:\vitor energia\FATURAS AGRICOLA 2025\2025\TODAS"  # Altere para o caminho da sua pasta
coordenadas = [(21.7, 361.2), (444.1, 571.7)]
output_xlsx = "resultado_final.xlsx"

# Processar todos os PDFs da pasta
df_resultado = processar_pasta_pdfs(pasta_pdfs, coordenadas, output_xlsx)

# Mostrar preview dos dados
print("\nPreview dos dados:")
print("=" * 80)
print(f"Total de arquivos processados: {len(df_resultado)}")
print(f"Colunas encontradas: {len(df_resultado.columns)}")
print(f"\nPrimeiras linhas:")
print(df_resultado.head())